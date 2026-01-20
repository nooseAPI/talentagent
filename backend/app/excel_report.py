from openpyxl import Workbook
from pathlib import Path
import pandas as pd

BASE_DIR = Path(__file__).resolve().parent
REPORT_DIR = BASE_DIR / "reports"
REPORT_DIR.mkdir(exist_ok=True)


def create_excel_report(project_name: str, body_text: str):

    file_path = REPORT_DIR / f"weekly_report_{project_name}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "PMO Weekly Report"

    ws["A1"] = f"프로젝트: {project_name}"

    row = 3
    for line in body_text.splitlines():
        ws[f"A{row}"] = line
        row += 1

    wb.save(file_path)

    return file_path


def excel_to_text(file_path: str) -> list[str]:
    df = pd.read_excel(file_path)

    texts = []
    for _, row in df.iterrows():
        text = " | ".join([f"{col}: {row[col]}" for col in df.columns])
        texts.append(text)

    return texts